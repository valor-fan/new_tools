# -*- coding: utf-8 -*-
import base64
import random
import requests
import os
import warnings
from openpyxl import load_workbook as load_xlsx_workbook, Workbook
from openpyxl.utils import get_column_letter
from termcolor import cprint
import time

# 忽略无关警告
requests.packages.urllib3.disable_warnings()
warnings.filterwarnings('ignore')
os.environ["TF_CPP_MIN_LOG_LEVEL"] = '1'

# ===================== Hunter 配置 =====================
config = {
    'api_key': '',
    'account': '',
    'size_threshold': '10',
    'search_url': '',
    'page': 1,
    'page_size': 20,
    'port_filter': False,
    'sleep_range': (15, 25)
}

# 企查查 Excel 文件配置
EXCEL_FILE = '1.xlsx'

# 基础过滤关键词
BANNED_WORDS = ['手机免费在线观看', '伦理', '高清完整版', '成人视频', '香蕉娱乐', '威尼斯']

# 新增：不良/非法内容过滤关键词
HARMFUL_CONTENT_WORDS = [
    "真人视讯", "视讯", "彩票", "彩票投注", "棋牌", "棋牌游戏",
    "赌博", "赌博网站", "赌场", "网赌", "博彩", "电竞投注",
    "捕鱼", "打鱼", "老虎机", "百家乐", "牛牛", "炸金花"
]

# 用于第二次筛查的关键词列表
SENSITIVE_SYSTEM_KEYWORDS = [
    "登录", "管理系统", "后台管理", "后台", "控制系统",
    "系统", "平台", "admin", "login", "manage",
    "control", "system", "platform"
]

# 结果文件名称
RESULTS_TXT_FILE = 'hunter_results.txt'
RESULTS_XLSX_FILE = 'hunter_results.xlsx'
SENSITIVE_SYSTEMS_XLSX_FILE = 'chinese_company_sensitive_systems.xlsx'
RESULTS_DIR = 'hunter_results'


def output(msg):
    """带时间戳的彩色日志输出"""
    color = "red" if "error" in msg.lower() else "green"
    time_now = time.strftime("%Y-%m-%d %X")
    cprint(f"[{time_now}]: {msg}", color)


def base64url_encode(s):
    """Hunter 标准 base64url 编码"""
    encoded = base64.urlsafe_b64encode(s.encode('utf-8')).decode('utf-8')
    return encoded.rstrip('=')


def contains_harmful_content(text):
    """检查文本是否包含不良内容"""
    if not text:
        return False
    for word in HARMFUL_CONTENT_WORDS:
        if word in text:
            return True
    return False


def test_hunter_permission():
    """用测试查询验证API-KEY权限"""
    output('开始测试 Hunter API 权限...')
    test_query = 'is_web=true && web.title="test"'
    encoded_query = base64url_encode(test_query)
    params = {
        'api-key': config['api_key'],
        'search': encoded_query,
        'page': 1,
        'page_size': config['page_size'],
        'port_filter': 'false'
    }
    try:
        resp = requests.get(config['search_url'], params=params, timeout=20, verify=False)
        if resp.status_code == 200:
            try:
                result = resp.json()
                if result.get('code') == 200:
                    output('Hunter API 权限验证成功！')
                    return True
                else:
                    output(
                        f'权限测试失败：API 返回错误（code={result.get("code")}），msg={result.get("message", "无描述")}')
                    return False
            except Exception as e:
                output(f'权限测试 JSON 解析失败：{str(e)}')
                return False
        else:
            output(f'权限测试失败：HTTP 状态码 {resp.status_code}')
            return False
    except Exception as e:
        output(f'权限测试异常：{str(e)}')
        return False


def read_qcc_excel(file_path):
    """读取企查查 Excel 文件"""
    output(f'开始读取 Excel 文件：{file_path}')
    if not os.path.exists(file_path):
        output(f'错误：文件 {file_path} 不存在！')
        exit()
    try:
        workbook = load_xlsx_workbook(file_path, data_only=True, read_only=True)
        worksheet = workbook.active
        output(f'成功加载工作表：{worksheet.title}（共 {worksheet.max_row} 行）')

        company_col = None
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, cell_value in enumerate(header_row):
            if cell_value and '公司名称' in str(cell_value):
                company_col = idx
                output(f'已找到公司名称列：第 {idx + 1} 列（{cell_value}）')
                break
        if company_col is None:
            output('错误：未找到“公司名称”列！')
            exit()

        valid_companies = []
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= company_col: continue
            company_name = str(row[company_col]).strip() if row[company_col] else ''
            if company_name and len(company_name) >= 2:
                valid_companies.append((row_num, company_name))
            if row_num % 1000 == 0:
                output(f'已读取 {row_num} 行，找到 {len(valid_companies)} 个有效公司名称')

        workbook.close()
        output(f'Excel 读取完成，共找到 {len(valid_companies)} 个有效公司名称')
        return valid_companies
    except Exception as e:
        output(f'Excel 读取错误：{str(e)}')
        exit()


def format_company_name(name):
    """格式化公司名称"""
    suffixes = ['有限公司', '有限责任公司', '股份有限公司', '技术', '科技', '集团', '控股', '投资', '发展']
    for suffix in suffixes:
        name = name.replace(suffix, '')
    return name.replace('(', '').replace(')', '').strip()


def init_result_files():
    """初始化所有结果文件"""
    if not os.path.exists(RESULTS_DIR):
        os.makedirs(RESULTS_DIR)

    # 初始化主查询结果文件 (TXT)
    with open(os.path.join(RESULTS_DIR, RESULTS_TXT_FILE), 'w', encoding='utf-8') as f:
        f.write('Hunter 批量查询结果\n')
        f.write(f'查询时间：{time.strftime("%Y-%m-%d %X")}\n')
        f.write(f'查询公司总数：{len(companies)}\n')
        f.write('=' * 80 + '\n\n')

    # 初始化主查询结果文件 (XLSX)
    wb = Workbook()
    ws = wb.active
    ws.title = "查询结果"
    headers = ['公司名称', '站点标题', '域名', 'IP地址', '端口', '状态码', '响应时间']
    ws.append(headers)
    for i, width in enumerate([20, 50, 25, 15, 10, 10, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    wb.save(os.path.join(RESULTS_DIR, RESULTS_XLSX_FILE))

    # 初始化敏感系统筛查结果文件 (XLSX)
    wb_systems = Workbook()
    ws_systems = wb_systems.active
    ws_systems.title = "敏感系统登录页面"
    system_headers = ['公司名称', '站点标题', '域名', 'IP地址', '端口', '状态码', '响应时间']
    ws_systems.append(system_headers)
    for i, width in enumerate([20, 50, 25, 15, 10, 10, 15], 1):
        ws_systems.column_dimensions[get_column_letter(i)].width = width
    wb_systems.save(os.path.join(RESULTS_DIR, SENSITIVE_SYSTEMS_XLSX_FILE))

    output('所有结果文件初始化完成！')


def save_to_main_results(company_name, assets):
    """保存主查询结果（包含不良内容过滤）"""
    # 写入 TXT
    with open(os.path.join(RESULTS_DIR, RESULTS_TXT_FILE), 'a+', encoding='utf-8') as txt_file:
        txt_file.write(f'===== 公司：{company_name} =====\n')
        txt_file.write('-' * 80 + '\n')

        # 写入 XLSX
        try:
            wb = load_xlsx_workbook(os.path.join(RESULTS_DIR, RESULTS_XLSX_FILE))
            ws = wb.active
        except Exception as e:
            output(f"警告：打开主 XLSX 文件失败！错误: {e}")
            return

        for asset in assets:
            title = asset.get('web_title', '').strip().replace('\n', '').replace('\r', '')

            # 1. 过滤不良内容
            if contains_harmful_content(title):
                output(f'[主查询] 过滤不良内容站点：{title}（{asset.get("domain", "")}）')
                continue

            # 2. 过滤无关站点
            if any(word in title for word in BANNED_WORDS):
                output(f'[主查询] 过滤无关站点：{title}（{asset.get("domain", "")}）')
                continue

            row_data = [
                company_name, title, asset.get('domain', ''), asset.get('ip', ''),
                asset.get('port', ''), asset.get('status_code', ''),
                f"{asset.get('response_time', '')}ms" if asset.get('response_time') else ''
            ]
            ws.append(row_data)
            output(f'[主查询] 保存资产：{title}（{row_data[2]} | {row_data[3]}:{row_data[4]}）')
            txt_file.write(
                f'网站标题：{row_data[1]}\n域名：{row_data[2]}\nIP地址：{row_data[3]}\n端口：{row_data[4]}\n状态码：{row_data[5]}\n响应时间：{row_data[6]}\n' + '-' * 80 + '\n')

        wb.save(os.path.join(RESULTS_DIR, RESULTS_XLSX_FILE))
        txt_file.write('\n' + '=' * 80 + '\n\n')


def save_to_systems_results(company_name, assets):
    """保存敏感系统筛查结果（包含不良内容过滤）"""
    if not assets:
        return

    try:
        wb = load_xlsx_workbook(os.path.join(RESULTS_DIR, SENSITIVE_SYSTEMS_XLSX_FILE))
        ws = wb.active
    except Exception as e:
        output(f"警告：打开敏感系统 XLSX 文件失败！错误: {e}")
        return

    for asset in assets:
        title = asset.get('web_title', '').strip().replace('\n', '').replace('\r', '')

        # 1. 过滤不良内容
        if contains_harmful_content(title):
            output(f'[敏感系统筛查] 过滤不良内容站点：{title}（{asset.get("domain", "")}）')
            continue

        # 2. 过滤无关站点
        if any(word in title for word in BANNED_WORDS):
            output(f'[敏感系统筛查] 过滤无关站点：{title}（{asset.get("domain", "")}）')
            continue

        row_data = [
            company_name, title, asset.get('domain', ''), asset.get('ip', ''),
            asset.get('port', ''), asset.get('status_code', ''),
            f"{asset.get('response_time', '')}ms" if asset.get('response_time') else ''
        ]
        ws.append(row_data)
        output(f'[敏感系统筛查] 发现并保存：{title}（{row_data[2]}）')

    wb.save(os.path.join(RESULTS_DIR, SENSITIVE_SYSTEMS_XLSX_FILE))


def search_hunter(company_name, formatted_name, row_num, idx, total):
    """第一次查询：获取公司所有相关站点"""
    query = f'is_web=true && (web.title="{formatted_name}" || web.body="{formatted_name}") && ip.country="CN"'
    output(f'\n[主查询] 【{idx}/{total}】正在查询公司：{company_name}（第 {row_num} 行）')
    output(f'[主查询] Hunter 查询语句：{query}')

    try:
        encoded_query = base64url_encode(query)
    except Exception as e:
        output(f'[主查询] 查询语句编码失败：{str(e)}')
        return []

    params = {'api-key': config['api_key'], 'search': encoded_query, 'page': config['page'],
              'page_size': config['page_size'], 'port_filter': str(config['port_filter']).lower()}
    time.sleep(random.randint(*config['sleep_range']))

    try:
        response = requests.get(config['search_url'], params=params, timeout=30, verify=False)
        if response.status_code != 200:
            output(f'[主查询] 查询失败。状态码: {response.status_code}')
            return []
        result = response.json()
        if result.get('code') != 200:
            output(f'[主查询] 查询无结果或API错误: {result.get("message", "无描述")}')
            return []

        total_count = result.get('data', {}).get('total', 0)
        assets = result.get('data', {}).get('arr', [])
        output(f'[主查询] 查询成功：找到 {total_count} 条资产')
        if total_count > int(config['size_threshold']):
            save_to_main_results(company_name, assets)
        return assets
    except Exception as e:
        output(f'[主查询] 查询异常：{str(e)}')
        return []


def search_for_sensitive_systems(company_name, formatted_name, row_num, idx, total):
    """第二次筛查：专门查询该公司的敏感系统和登录页面（含IP去重和内容过滤）"""
    output(f'\n[敏感系统筛查] 【{idx}/{total}】正在为公司 {company_name} 筛查敏感系统...')

    all_sensitive_assets = []
    max_or_per_query = 5

    keyword_groups = [SENSITIVE_SYSTEM_KEYWORDS[i:i + max_or_per_query] for i in
                      range(0, len(SENSITIVE_SYSTEM_KEYWORDS), max_or_per_query)]

    for i, group in enumerate(keyword_groups):
        if not group: continue
        keywords_query = " || ".join([f'web.title="{kw}"' for kw in group])
        query = f'is_web=true && ((web.title="{formatted_name}" || web.body="{formatted_name}") && ({keywords_query})) && ip.country="CN"'

        output(f'[敏感系统筛查] 正在执行子查询 {i + 1}/{len(keyword_groups)}，关键词：{", ".join(group)}')

        try:
            encoded_query = base64url_encode(query)
        except Exception as e:
            output(f'[敏感系统筛查] 子查询语句编码失败：{str(e)}')
            continue

        params = {'api-key': config['api_key'], 'search': encoded_query, 'page': config['page'],
                  'page_size': config['page_size'], 'port_filter': str(config['port_filter']).lower()}
        time.sleep(random.randint(*config['sleep_range']))

        try:
            response = requests.get(config['search_url'], params=params, timeout=30, verify=False)
            if response.status_code != 200 or response.json().get('code') != 200:
                output(f'[敏感系统筛查] 子查询无结果或失败。')
                continue

            result = response.json()
            assets = result.get('data', {}).get('arr', [])
            output(f'[敏感系统筛查] 子查询成功：找到 {len(assets)} 条敏感资产')
            if assets:
                all_sensitive_assets.extend(assets)
        except Exception as e:
            output(f'[敏感系统筛查] 子查询异常：{str(e)}')

    # --- 核心增强：结果处理 ---
    unique_assets = []
    seen_urls = set()
    seen_ips = set()  # 用于存储已见过的IP地址

    for asset in all_sensitive_assets:
        ip_address = asset.get('ip', '')
        url = asset.get('url', '')

        # 1. 按IP去重
        if ip_address and ip_address in seen_ips:
            output(f'[敏感系统筛查] IP已存在，跳过重复项：{ip_address} ({asset.get("web_title", "")})')
            continue

        # 2. 按URL去重
        if url and url in seen_urls:
            continue

        # 3. 记录已见过的IP和URL
        if ip_address: seen_ips.add(ip_address)
        if url: seen_urls.add(url)

        unique_assets.append(asset)

    # --- 统一保存最终结果 ---
    if unique_assets:
        output(
            f'[敏感系统筛查] 筛查完成：为 {company_name} 共找到 {len(unique_assets)} 个不重复的疑似敏感系统/登录页面。')
        save_to_systems_results(company_name, unique_assets)
    else:
        output(f'[敏感系统筛查] 筛查完成：未找到 {company_name} 的有效敏感系统/登录页面。')


def main():
    global companies
    output('=' * 50)
    output('        企查查公司资产 Hunter 批量查询工具 (增强版)        ')
    output('=' * 50)

    if not test_hunter_permission():
        output('错误：Hunter API 权限验证失败，无法继续查询！')
        return

    companies = read_qcc_excel(EXCEL_FILE)
    if not companies:
        output('错误：未找到有效公司名称！')
        return

    init_result_files()

    total = len(companies)
    try:
        for idx, (row_num, company_name) in enumerate(companies, start=1):
            formatted_name = format_company_name(company_name)

            # 1. 第一次查询
            search_hunter(company_name, formatted_name, row_num, idx, total)

            # 2. 第二次筛查
            search_for_sensitive_systems(company_name, formatted_name, row_num, idx, total)

            output(f"\n--- 公司【{idx}/{total}】{company_name} 的两次查询流程全部完成 ---")

        output('\n' + '=' * 50)
        output(f'所有 {total} 家公司的查询和筛查流程全部完成！')
        output(f'主查询结果保存在: {os.path.join(RESULTS_DIR, RESULTS_XLSX_FILE)}')
        output(f'敏感系统筛查结果保存在: {os.path.join(RESULTS_DIR, SENSITIVE_SYSTEMS_XLSX_FILE)}')
        output('=' * 50)
    except KeyboardInterrupt:
        output('\n' + '=' * 50)
        output(f'操作被手动中断！已完成 {idx - 1}/{total} 家公司的全部流程。')
        output('=' * 50)
    except Exception as e:
        output(f'\n全局异常：{str(e)}')
        output(f'已完成 {idx - 1}/{total} 家公司的全部流程。')


if __name__ == '__main__':
    main()